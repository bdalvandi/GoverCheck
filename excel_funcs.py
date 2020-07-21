from openpyxl import load_workbook


class XlsmWrite:
    def __init__(self, filename):
        self.wb = load_workbook(filename)
        self.sh = None

    def WriteRowTuple(self, sheet, row, start_col, tup):
        self.sh = self.wb[sheet]
        for col, value in enumerate(tup):
            self.sh.cell(row,start_col+col).value = value

    def WriteCell(self, sheet, row, col, value):
        self.sh = self.wb[sheet]
        self.sh.cell(row, col).value = value

    def Save(self, filename):
        self.wb.save(filename)
