# public functions used in project

from openpyxl import load_workbook

# shift list "L" elements to left and repeat last element "n" times
def shift(L, n): # list
    ext = [L[-1]] * len(L)
    return (L[n:] + ext)[:len(L)]


# get number of rows from openpyxl.sheet.calculate_dimension as str
def get_row_count(coord): # str
    i = -1
    rc = ''
    while True:
        if coord[i].isnumeric():
            rc = coord[i] + rc
            i -= 1
        else:
            return rc


# load droop results from last __trend.xlsx file
def droop_results(file_name):
    wb = load_workbook(file_name, read_only=True, data_only=True)
    sh = wb['Droop']
    row_count = int(get_row_count(sh.calculate_dimension()))
    drp_dic = dict()
    for r in range(2, row_count):
        drp_dic[sh.cell(r, 1).value] = sh.cell(r, 11).value

    del wb, sh
    return drp_dic


# class for trend values from "__trend.xlsm" and make data for individual excel files
class IndividualTrends:
    def __init__(self, sheet, range):
        self.sh = sheet
        self.rng = range

    @property
    def TimeValues(self):
        return [self.sh.cell(1, c).value for c in range(10, 677)]

    @property
    def FrqValues(self):
        return [self.sh.cell(2, c).value for c in range(10, 677)]

    @property
    def PowerValues(self):
        pwr_dic = dict()
        for row in self.rng[2:]:
            pwr_values = []
            n = int()
            unit = str()
            for c, cell in enumerate(row):
                if c > 668:
                    break
                if c == 0:
                    n = cell.value
                elif c == 1:
                    unit = cell.value
                else:
                    pwr_values.append(cell.value)
            pwr_values = shift(pwr_values, n)
            pwr_dic[unit] = pwr_values

        return pwr_dic


class GenerateExcelChart:
    def __init__(self, file_name):
        self.fname = file_name

    def __enter__(self):
        self.wb = load_workbook('__chart.xlsx')
        return self.wb

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(self.fname)
        del self.wb
