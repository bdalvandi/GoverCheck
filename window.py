from _window import mainWin
from getpmr import Pmr
from openpyxl import load_workbook
from project_functions import *
import wx
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import os, shutil, glob


class MainWin(mainWin):
    def __init__(self, parent):
        super().__init__(parent)
        self.pmr = None
        self.frq = None
        self.time = None


    def onOpenPMR(self, event):
        with wx.FileDialog(self, "Open PMR file", wildcard="PMR files (*.RPM)|*.RPM",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return  # the user changed their mind

            # Proceed loading the file chosen by the user
            self.txtPmrFilename.SetValue(dlg.GetPath())


        # extract PMR data
        wait = wx.BusyCursor()
        self.pmr = Pmr(self.txtPmrFilename.Value)
        del wait
        self.txtDate.SetValue(self.pmr.eventDate)
        self.txtTime.SetValue(self.pmr.eventTime)

        # enable next controls
        self.btnPreview.Enable()

    def onPreviewFreq(self, event):
        # get frq trend from pmr
        trend_dic = self.pmr.dataAsDict
        frq = trend_dic['3058FRHZ']
        self.time = tuple(frq.keys())
        self.frq = tuple(frq.values())

        # fill time dropdown choices
        t_list = tuple(map(lambda x: str(x)[-8:], self.time))
        self.selStart.SetItems(t_list)
        self.selPeak.SetItems(t_list)
        self.selEnd.SetItems(t_list)

        # show frq trend
        fig, ax = plt.subplots()
        fig.canvas.set_window_title('Preview Frequency Curve')
        fig.suptitle('Frequency')
        ax.plot(self.time, self.frq)
        fmt = mdates.DateFormatter('%H:%M:%S')
        ax.xaxis.set_major_formatter(fmt)
        ax.set_ylabel('Hz')
        plt.show()

        # enable next controls
        self.btnCreateTrend.Enable()


    def onCreateTrends(self, event):
        # make temp absolute paths
        tmp = 'c:/_gctmp'
        if os.path.isdir(tmp):
            files = filter(os.path.isfile, glob.glob(tmp + '/*.*'))
            for fl in files:
                os.remove(fl)
        else:
            os.mkdir(tmp)
        shutil.copyfile('__trends.xlsm', tmp + '/__trends.xlsm')

        # generate pmr data as text
        pmr_dic = self.pmr.dataAsDict
        with open(tmp + '/__testresult.CSV', 'w') as f:
            frq_dic = pmr_dic['3058FRHZ']
            f.write(',')
            f.write(','.join(map(lambda x: str(x)[-8:], frq_dic.keys())))
            f.write('\nFrequency,')
            f.write(','.join(map(str, frq_dic.values())))
            f.write('\n')
            for key in pmr_dic:
                if key == '3058FRHZ':
                    continue
                t_v = pmr_dic[key]
                f.write(self.pmr.unitsAsDict[key][0] + ',')
                f.write(','.join(map(str, t_v.values())))
                f.write('\n')

        with open(tmp + '//__testtimes.CSV', 'w') as f:
            t_start = self.selStart.GetString(self.selStart.GetCurrentSelection())
            t_peak = self.selPeak.GetString(self.selPeak.GetCurrentSelection())
            t_end = self.selEnd.GetString(self.selEnd.GetCurrentSelection())
            t_date = self.txtDate.GetValue()
            t_time = self.txtTime.GetValue()

            times = ','.join((t_start, t_peak, t_end, '', t_date, t_time))
            f.write(times)

        # open trends chart
        os.system(r'c:\_gctmp\__trends.xlsm')

        # enable next controls
        self.btnShowTrend.Enable()
        self.btnSelectFolder.Enable()
        self.txtFolder.Enable()
        # self.chkAddDate.Enable()
        self.selTag.Enable()
        self.btnGen.Enable()

    def onShowTrends(self, event):
        os.system(r'c:\_gctmp\__trends.xlsm')

    def onSelFolder(self, event):
        with wx.DirDialog(self, 'Select the test result folder') as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return

            self.txtFolder.SetValue(dlg.GetPath())

    def onGenerate(self, event):
        try:
            wait = wx.BusyCursor()
            shutil.copyfile(r'c:\_gctmp\__trends.xlsm', r'c:\_gctmp\__trends_cpy.xlsm')
            wb = load_workbook(r'c:\_gctmp\__trends_cpy.xlsm')
            sh = wb['Data']

            # get row count :rc
            rc = get_row_count(sh.calculate_dimension())
            rng = sh['H1':'ZZ' + rc]
            tr = IndividualTrends(sh, rng)
            del wb, sh

            droop = droop_results(r'c:\_gctmp\__trends.xlsm')

            # file name date & tag
            fname_tag = self.txtDate.Value
            fname_tag = '_' + fname_tag.replace('-', '')
            fname_tag += self.selTag.GetString(self.selTag.GetCurrentSelection())

            # writing data to sheet
            for i, unit in enumerate(tr.PowerValues):
                # self.progress.SetValue(i * 100 // len(tr.PowerValues))
                with GenerateExcelChart(self.txtFolder.GetValue() + '/' + unit + fname_tag + '.xlsx') as wb:
                    wb.iso_dates = True
                    sh = wb['Data']

                    for c, t in enumerate(tr.TimeValues, 2):
                        sh.cell(1, c).value = t
                    for c, f in enumerate(tr.FrqValues, 2):
                        sh.cell(2, c).value = f
                    for c, p in enumerate(tr.PowerValues[unit], 2):
                        sh.cell(3, c).value = p
                    try:
                        drp_val = '\tDroop: {:.1%}\n'.format(droop[unit])
                    except ValueError:
                        drp_val = '\tNo Droop\n'
                    sh['A5'] = unit.replace('_GEN_', '_') + drp_val + self.txtDate.Value + '\t' + self.txtTime.Value
        finally:
            del wait
